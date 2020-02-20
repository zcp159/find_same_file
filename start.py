import hashlib
import fnmatch
import os
import openpyxl
import threading
import queue
import pyautogui
import time

dic = {}  # 放置{"MD5":["相同MD5文件路径1","相同MD5文件路径2","..."]}
q = queue.Queue()  # 安全队列用于放置所有文件路径
threads = []  # 放置所有线程
is_jisuan_size ="y" # 放置是否计算文件大小

# 确定文件是否符合规则
def is_file_math(file, guizes):
    # 遍历规则，如果文件符合任何一条规则则返回真
    for guize in guizes:
        if fnmatch.fnmatch(file, guize):
            return True
    # 遍历完成后，返回假
    # 如果前面已经返回真了，由于只能返回一个值，则这个返回假就没用了
    return False


# 获取目录及规则，返回目录下所有满足规则的文件的绝对路径
def is_special_file(root, guizes=["*"], liwaimulus=[]):
    # 遍历根目录
    for dangqianmulu, mulus, files in os.walk(os.path.abspath(root)):
        # 遍历返回的文件名
        for file in files:
            # 判定文件名是否满足需要搜索的格式,如果符合则绝对路径放入生成器
            if is_file_math(file, guizes):
                yield os.path.join(os.path.abspath(dangqianmulu), file)
        for d in liwaimulus:
            if d in mulus:
                mulus.remove(d)


# 计算文件MD5
def get_file_md5(file):
    d = hashlib.md5()
    for i in split_file(file):
        d.update(i)
    return d.hexdigest()


# 将传过来的文件名读取，分片后返回。
def split_file(file):
    with open(file, "rb")as f:
        while True:
            s = f.read(8192)
            if not s:
                break
            else:
                yield s


# 接收文件名，构建该文件和py文件在一起的绝对路径
def build_pymulu_filename(file_name):
    # 先得到py文件的目录，再加上文件名构建路径
    dangqianmulu = os.path.split(os.path.abspath(__file__))[0]
    return os.path.join(dangqianmulu, file_name)


# 传递过来文件绝对路径，计算MD5后添加进字典{"MD5":["相同MD5文件路径1","相同MD5文件路径2","..."]}
def MD5jisuan_dic():
    global dic
    global is_jisuan_MD5
    while True:
        name = q.get()
        if name is None:
            break
        # 如果生成键值对不存在，则新建{"MD5":[]}
        # 把有相同MD5的文件放入一个键值对内，值为字符串列表
        current_md5 = get_file_md5(name)
        if not name == dic.setdefault(current_md5, []):
            dic[current_md5].append(name)
        # # 如果列表长度为1，则是第一个MD5文件，计算文件大小，时间多了6倍
        if len(dic[current_md5]) == 1 and is_jisuan_size == "y":
            dic[current_md5].insert(0, int(os.path.getsize(name) / 1024 / 1024))
        q.task_done()


def main():
    # 用户输入路径
    m = pyautogui.prompt('请输入目录，如输入: D:\Program Files\Python ')
    if str(m) == "None":
        exit()
    # 用户确认是否计算文件大小
    is_jisuan_size = pyautogui.prompt('是否计算文件大小，如计算时间多6倍，输入小写y/n ')
    if is_jisuan_size != "y" and is_jisuan_size != "n":
        exit()

    # 计时开始
    time_start = time.perf_counter()

    # 新建EXCEL
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'MD5'
    ws['B1'] = '文件大小(M)'
    ws['C1'] = '文件1'
    ws['D1'] = '文件2'
    ws['E1'] = '文件3'
    ws['F1'] = '文件4'

    # 得到目录下所有满足条件的文件绝对路径
    lujing_all = is_special_file(m)

    # 启动10个线程，计算MD5
    for i in range(10):
        t = threading.Thread(target=MD5jisuan_dic)
        t.start()
        threads.append(t)

    # 把所有文件路径放入队列
    for item in lujing_all:
        q.put(item)

    # 等待队列消费完毕
    q.join()

    # 停止队列
    for i in range(10):
        q.put(None)
    # 停止线程的无限循环
    for t in threads:
        t.join()

    # 遍历文件MD5字典
    j = 2  # 行指针
    for i in dic:
        # 输出key，key对应字典键值
        print(i, dic[i])
        # 把MD5写进EXCEL首列
        ws.cell(row=j, column=1).value = i
        # 遍历相同MD5的文件名们
        for n in range(0, len(dic[i])):
            # 把文件名填入EXCEL，与MD5对应的行
            ws.cell(row=j, column=n + 2).value = dic[i][n]
        j = j + 1

    # 保存到py文件目录下
    wb.save(build_pymulu_filename("jisuanMD5.xlsx"))
    wb.close()
    print("结束 {}".format(time.perf_counter() - time_start))


if __name__ == '__main__':
    main()
